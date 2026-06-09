import argparse
import re
import csv
import datetime
import math
import os
from collections import defaultdict

import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

from aux_functions import get_date_format, get_trades_col_names, get_dividends_col_names
from exchange_rates import ExchangeRateProvider


# ---------------------------------------------------------------------------
# Styling constants
# ---------------------------------------------------------------------------
_H1_FILL        = PatternFill(start_color='1F6B3A', end_color='1F6B3A', fill_type='solid')
_H2_FILL        = PatternFill(start_color='1F3E6B', end_color='1F3E6B', fill_type='solid')
_H1_TOTALS_FILL = PatternFill(start_color='E8F4EA', end_color='E8F4EA', fill_type='solid')
_H2_TOTALS_FILL = PatternFill(start_color='E8EEF8', end_color='E8EEF8', fill_type='solid')
_HEADER_FONT    = Font(bold=True, color='FFFFFF', size=12)
_TOTALS_FONT    = Font(bold=True, size=11)
_TABLE_GAP      = 1
_THIN_BORDER    = Border(left=Side(style='thin'), right=Side(style='thin'),
                         top=Side(style='thin'),  bottom=Side(style='thin'))


def round_half_up(x):
    return math.floor(x + 0.5) if x >= 0 else math.ceil(x - 0.5)


def _style_header_row(sheet, row, col_start, col_end, label, fill):
    for col in range(col_start, col_end + 1):
        cell = sheet.cell(row=row, column=col)
        cell.fill = fill
        cell.font = _HEADER_FONT
        cell.alignment = Alignment(horizontal='center', vertical='center')
    label_cell = sheet.cell(row=row, column=col_start)
    label_cell.value = label
    label_cell.alignment = Alignment(horizontal='left', vertical='center')
    sheet.row_dimensions[row].height = 22


def _style_totals_row(sheet, row, col_start, col_end, fill):
    for col in range(col_start, col_end + 1):
        cell = sheet.cell(row=row, column=col)
        cell.fill = fill
        cell.font = _TOTALS_FONT


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------
def parse_number(s):
    return float(s.replace(',', ''))


def get_row_number(row, col_names, col_name, default=0):
    if col_name not in col_names:
        return default
    value = row[col_names[col_name]]
    return default if value == '' else parse_number(value)


def get_multiplier(asset_category):
    return 100 if asset_category == 'Equity and Index Options' else 1


def _in_half(dt, half_label):
    return (dt.month <= 6) if half_label == 'H1' else (dt.month > 6)


def _csv_path(file_dir, csv_file_name):
    return f'{file_dir}/{csv_file_name}.csv'


def calculate_taxable_profit(profit, profit_trivial, profit_adjusted):
    if profit >= 0:
        return max(min(profit_trivial, profit_adjusted), 0)
    return min(max(profit_trivial, profit_adjusted), 0)


# ---------------------------------------------------------------------------
# Core trade logic
# ---------------------------------------------------------------------------
def resolve_opening_info(closed_lot_trades, opening_orders, multiplier):
    """
    Match each ClosedLot to opening Order rows.
    Returns list of {'open_price', 'allocated_open_fee'} or None per lot.
    """
    if not opening_orders:
        return [None] * len(closed_lot_trades)

    remaining = [{**o, 'remaining': abs(o['quantity'])} for o in opening_orders]

    results = []
    for lot in closed_lot_trades:
        lot_qty  = abs(lot['quantity'])
        lot_date = lot['datetime'].date()

        candidates = [r for r in remaining
                      if r['ticker'] == lot['ticker']
                      and r['datetime'].date() == lot_date
                      and r['remaining'] > 0]

        if not candidates:
            results.append(None)
            continue

        lot_basis      = abs(lot['basis'])
        matched_single = None
        for c in candidates:
            fee_per_share  = abs(c['fee']) / abs(c['quantity'])
            sign           = 1 if lot['quantity'] > 0 else -1
            expected_basis = (abs(c['price']) * multiplier + sign * fee_per_share) * lot_qty
            if abs(expected_basis - lot_basis) < 0.005:
                matched_single = c
                break

        if matched_single:
            matched_single['remaining'] -= lot_qty
            results.append({
                'open_price':        abs(matched_single['price']),
                'allocated_open_fee': abs(matched_single['fee']) * lot_qty / abs(matched_single['quantity']),
            })
        else:
            weighted_price, total_fee, qty_left = 0.0, 0.0, lot_qty
            for c in candidates:
                take = min(c['remaining'], qty_left)
                if take <= 0:
                    continue
                weighted_price  += abs(c['price']) * take
                total_fee       += abs(c['fee']) * take / abs(c['quantity'])
                c['remaining']  -= take
                qty_left        -= take
                if qty_left <= 0:
                    break
            results.append(None if qty_left > 0 else {
                'open_price':        weighted_price / lot_qty,
                'allocated_open_fee': total_fee,
            })

    return results


def build_closed_lot_dict(closed_lot_trade, closing_trade, allocated_close_fee, rate_provider, opening_resolution=None):
    multiplier    = get_multiplier(closed_lot_trade['asset_category'])
    lot_quantity  = closed_lot_trade['quantity']
    abs_quantity  = abs(lot_quantity)
    is_long       = lot_quantity > 0
    position_type = 'long' if is_long else 'short'

    open_price_resolved = opening_resolution['open_price']         if opening_resolution else None
    open_fee_resolved   = opening_resolution['allocated_open_fee'] if opening_resolution else None

    opening_basis     = closed_lot_trade['basis'] or lot_quantity * closed_lot_trade['price'] * multiplier
    close_gross_value = abs_quantity * closing_trade['price'] * multiplier
    form_open_dt      = closed_lot_trade['datetime']
    tax_event_dt      = closing_trade['datetime']

    if is_long:
        original_value      = abs(opening_basis)
        consideration_value = close_gross_value + allocated_close_fee
        profit_trivial_rates = (consideration_value, tax_event_dt, original_value, form_open_dt)
    else:
        open_price_clean    = open_price_resolved if open_price_resolved is not None else closed_lot_trade['price']
        open_gross_value    = abs_quantity * open_price_clean * multiplier
        original_value      = close_gross_value - allocated_close_fee
        consideration_value = abs(opening_basis)
        profit_trivial_rates = (consideration_value, form_open_dt, original_value, tax_event_dt)

    cur = closed_lot_trade['currency']
    rate = lambda dt: rate_provider.get_rate(cur, dt)

    open_currency_factor          = rate(form_open_dt)
    close_currency_factor         = rate(tax_event_dt)
    currency_factor_ratio         = close_currency_factor / open_currency_factor if open_currency_factor != 0 else 0
    open_value_ILS                = original_value * open_currency_factor
    open_value_ILS_adjusted_forex = open_value_ILS * currency_factor_ratio
    close_value_ILS               = consideration_value * close_currency_factor

    profit = consideration_value - original_value
    tc, tc_dt, to, to_dt = profit_trivial_rates
    profit_trivial_ILS    = tc * rate(tc_dt) - to * rate(to_dt)
    profit_adjusted_forex = close_value_ILS - open_value_ILS_adjusted_forex

    print(f"Ticker: {closed_lot_trade['ticker']}")
    return {
        'currency':                      cur,
        'ticker':                        closed_lot_trade['ticker'],
        'tax_event_datetime':            tax_event_dt,
        'form_open_datetime':            form_open_dt,
        'open_date':                     form_open_dt.strftime("%d/%m/%Y"),
        'close_date':                    tax_event_dt.strftime("%d/%m/%Y"),
        'quantity':                      lot_quantity,
        'position_type':                 position_type,
        'open_price':                    closing_trade['price'] if not is_long else (open_price_resolved or closed_lot_trade['price']),
        'close_price':                   (open_price_resolved or closed_lot_trade['price']) if not is_long else closing_trade['price'],
        'allocated_open_fee':            open_fee_resolved or 0,
        'allocated_close_fee':           allocated_close_fee,
        'open_value':                    original_value,
        'close_value':                   consideration_value,
        'open_currency_factor':          open_currency_factor,
        'close_currency_factor':         close_currency_factor,
        'currency_factor_ratio':         currency_factor_ratio,
        'open_value_ILS':                open_value_ILS,
        'open_value_ILS_adjusted_forex': open_value_ILS_adjusted_forex,
        'close_value_ILS':               close_value_ILS,
        'gross_sale_value_ILS':          open_gross_value * open_currency_factor if not is_long else close_gross_value * close_currency_factor,
        'profit':                        profit,
        'profit_trivial_ILS':            profit_trivial_ILS,
        'profit_ILS_forex':              calculate_taxable_profit(profit, profit_trivial_ILS, profit_adjusted_forex),
    }


def parse_trade_row(row, col_names, date_slash_format):
    dt_str = row[col_names['datetime']]
    return {
        'trade_type':     row[col_names['trade_type']],
        'asset_category': row[col_names['asset_category']],
        'currency':       row[col_names['currency']],
        'ticker':         row[col_names['ticker']],
        'datetime':       datetime.datetime.strptime(dt_str, get_date_format(dt_str, date_slash_format=date_slash_format)),
        'quantity':       parse_number(row[col_names['quantity']]),
        'price':          get_row_number(row, col_names, 'price'),
        'fee':            get_row_number(row, col_names, 'fee'),
        'basis':          get_row_number(row, col_names, 'basis'),
    }


# ---------------------------------------------------------------------------
# Extractors
# ---------------------------------------------------------------------------
def extract_trades_data_from_csv(file_dir, csv_file_name, verbosity=0, date_slash_format='normal'):
    csv_file      = _csv_path(file_dir, csv_file_name)
    rate_provider = ExchangeRateProvider()
    col_names     = get_trades_col_names(csv_file)

    closed_lots_list          = []
    closed_lots_datetime_list = []
    inds_sorted_close_dates   = []

    if not col_names:
        print('no trades exist in the file.')
        return closed_lots_list, inds_sorted_close_dates

    # Pass 1: collect closing trades and opening orders
    closing_trade_pairs   = []
    opening_orders        = []
    current_closing_trade = None
    current_closed_lots   = []

    with open(csv_file, 'r') as f:
        for row in csv.reader(f):
            if verbosity == 1:
                print(row)
            if (row[col_names['main']] != 'Trades' or row[col_names['header']] != 'Data'
                    or row[col_names['asset_category']] not in ('Stocks', 'Equity and Index Options')):
                continue
            trade_type = row[col_names['trade_type']]
            open_close = row[col_names['code']] if 'code' in col_names else ''
            is_open_only = 'O' in open_close and 'C' not in open_close

            if trade_type == 'Order' and is_open_only:
                p = parse_trade_row(row, col_names, date_slash_format)
                opening_orders.append({k: p[k] for k in ('ticker', 'datetime', 'price', 'fee', 'quantity')})
            elif trade_type == 'Trade' and not is_open_only:
                if current_closing_trade is not None:
                    closing_trade_pairs.append((current_closing_trade, current_closed_lots))
                current_closing_trade = parse_trade_row(row, col_names, date_slash_format)
                current_closed_lots   = []
            elif 'ClosedLot' in trade_type:
                current_closed_lots.append(parse_trade_row(row, col_names, date_slash_format))

    if current_closing_trade is not None:
        closing_trade_pairs.append((current_closing_trade, current_closed_lots))

    # Pass 2: group lots by ticker, resolve opening info per ticker
    ticker_lots = defaultdict(list)
    for closing_trade, lots in closing_trade_pairs:
        total_qty = sum(abs(t['quantity']) for t in lots)
        for lot in lots:
            fee = closing_trade['fee'] * abs(lot['quantity']) / total_qty if total_qty else 0
            ticker_lots[lot['ticker']].append((lot, closing_trade, fee))

    ticker_resolutions = {
        ticker: resolve_opening_info(
            [t[0] for t in tuples],
            [o for o in opening_orders if o['ticker'] == ticker],
            get_multiplier(tuples[0][0]['asset_category']),
        )
        for ticker, tuples in ticker_lots.items()
    }

    # Pass 3: build closed_lot_dicts in CSV order
    ticker_idx = defaultdict(int)
    for closing_trade, lots in closing_trade_pairs:
        total_qty = sum(abs(t['quantity']) for t in lots)
        for lot in lots:
            fee        = closing_trade['fee'] * abs(lot['quantity']) / total_qty if total_qty else 0
            ticker     = lot['ticker']
            resolution = ticker_resolutions[ticker][ticker_idx[ticker]]
            ticker_idx[ticker] += 1
            d = build_closed_lot_dict(lot, closing_trade, fee, rate_provider, resolution)
            closed_lots_list.append(d)
            closed_lots_datetime_list.append(d['tax_event_datetime'])

    if verbosity == 1:
        for d in closed_lots_list:
            print(f"ticker {d['ticker']}: open={d['open_date']} close={d['close_date']} "
                  f"type={d['position_type']} qty={d['quantity']} profit={d['profit']} "
                  f"open_rate={d['open_currency_factor']} close_rate={d['close_currency_factor']}")

    inds_sorted_close_dates = sorted(range(len(closed_lots_datetime_list)),
                                     key=lambda i: closed_lots_datetime_list[i])
    return closed_lots_list, inds_sorted_close_dates


def extract_dividends_data_from_csv(file_dir, csv_file_name, verbosity=0, date_slash_format='normal'):
    """
    Single-pass extraction. Returns (dividends_list, unmatched_wht).
    """
    csv_file      = _csv_path(file_dir, csv_file_name)
    rate_provider = ExchangeRateProvider()
    col_names     = get_dividends_col_names(csv_file)

    if not col_names:
        print('no dividends exist in the file.')
        return [], []

    dividends_list = []
    all_wht_rows   = []

    with open(csv_file, 'r') as f:
        for row in csv.reader(f):
            if verbosity == 1:
                print(row)
            if row[col_names['main']] not in ('Dividends', 'Withholding Tax') or row[col_names['header']] != 'Data':
                continue
            currency = row[col_names['currency']]
            if 'Total' in currency:
                continue
            dt_str   = row[col_names['datetime']]
            dt       = datetime.datetime.strptime(dt_str, get_date_format(dt_str, date_slash_format=date_slash_format))
            date_str = dt.strftime("%d/%m/%Y")
            ticker   = row[col_names['ticker']].split('(')[0]
            amount   = float(row[col_names['amount']])

            if row[col_names['main']] == 'Dividends':
                if dividends_list and dividends_list[-1]['ticker'] == ticker and dividends_list[-1]['date'] == date_str:
                    dividends_list[-1]['amount'] += amount
                else:
                    dividends_list.append({'currency': currency, 'datetime': dt, 'date': date_str,
                                           'ticker': ticker, 'amount': amount, 'withholding_tax': 0})
            else:
                all_wht_rows.append({'ticker': ticker, 'date': date_str, 'datetime': dt,
                                     'currency': currency, 'amount': amount})

    dividend_keys = {(d['ticker'], d['date']) for d in dividends_list}
    for wht in all_wht_rows:
        if (wht['ticker'], wht['date']) in dividend_keys:
            for d in dividends_list:
                if d['ticker'] == wht['ticker'] and d['date'] == wht['date']:
                    d['withholding_tax'] += wht['amount']
                    break

    for d in dividends_list:
        d['currency_factor']     = rate_provider.get_rate(d['currency'], d['datetime'])
        d['dividend']            = d['amount']
        d['dividend_ILS']        = d['dividend'] * d['currency_factor']
        d['withholding_tax_ILS'] = d['withholding_tax'] * d['currency_factor']

    unmatched_wht = []
    for wht in all_wht_rows:
        if (wht['ticker'], wht['date']) not in dividend_keys:
            rate = rate_provider.get_rate(wht['currency'], wht['datetime'])
            unmatched_wht.append({**wht, 'rate': rate, 'amount_ils': wht['amount'] * rate})

    return dividends_list, unmatched_wht


def extract_other_fees_data_from_csv(file_dir, csv_file_name, verbosity=0, date_slash_format='normal'):
    csv_file      = _csv_path(file_dir, csv_file_name)
    rate_provider = ExchangeRateProvider()
    result        = {'commissions': [], 'fees': [], 'interest_debit': [], 'interest_credit': []}

    def _make_entry(currency, dt_str, amount_str, description, category):
        dt     = datetime.datetime.strptime(dt_str, get_date_format(dt_str, date_slash_format=date_slash_format))
        amount = float(amount_str)
        if amount == 0:
            return None
        fx = rate_provider.get_rate(currency, dt)
        return {'date': dt.strftime('%d/%m/%Y'), 'description': description, 'currency': currency,
                'amount': amount, 'rate': fx, 'amount_ils': amount * fx, 'category': category}

    with open(csv_file, 'r') as f:
        for row in csv.reader(f):
            if len(row) < 4 or row[1].strip() != 'Data':
                continue
            section = row[0].strip()
            try:
                if section == 'Commission Details':
                    if len(row) < 8:
                        continue
                    symbol = row[4].strip()
                    parts  = symbol.split('.')
                    if len(parts) != 2 or not parts[0].isalpha() or not parts[1].isalpha() or not row[5].strip():
                        continue
                    entry = _make_entry(parts[0], row[5].strip(), row[7], symbol, 'Commission')
                    if entry:
                        result['commissions'].append(entry)

                elif section == 'Fees':
                    if len(row) < 7:
                        continue
                    currency = row[3].strip()
                    if 'Total' in currency or not currency or not row[4].strip():
                        continue
                    entry = _make_entry(currency, row[4].strip(), row[6], row[5].strip() if len(row) > 5 else '', 'Fee')
                    if entry:
                        result['fees'].append(entry)

                elif section == 'Interest':
                    if len(row) < 6:
                        continue
                    currency = row[2].strip()
                    if 'Total' in currency or not currency or not row[3].strip():
                        continue
                    description = row[4].strip() if len(row) > 4 else ''
                    entry = _make_entry(currency, row[3].strip(), row[5], description, 'Interest')
                    if entry:
                        if 'Borrow Fee' in description:
                            entry['category'] = 'Fee'
                            result['fees'].append(entry)
                        elif 'Debit Interest' in description:
                            result['interest_debit'].append(entry)
                        else:
                            result['interest_credit'].append(entry)
            except (ValueError, IndexError, KeyError):
                continue

    return result


# ---------------------------------------------------------------------------
# Excel writers
# ---------------------------------------------------------------------------
def _write_capital_gains_half(sheet, closed_lots_list, inds_sorted_close_dates, half_label, start_row):
    total_net, total_profit, total_loss, total_sell = 0, 0, 0, 0
    ind_line = 0

    for idx in inds_sorted_close_dates:
        d = closed_lots_list[idx]
        if not _in_half(d['tax_event_datetime'], half_label):
            continue
        r = start_row + ind_line
        sheet[f'A{r}'] = ind_line + 1
        sheet[f'B{r}'] = d['ticker']
        sheet[f'D{r}'] = d['currency']
        sheet[f'E{r}'] = d['close_value']
        sheet[f'F{r}'] = d['open_date']
        sheet[f'G{r}'] = d['open_value']
        sheet[f'H{r}'] = d['open_value_ILS']
        sheet[f'I{r}'] = d['open_currency_factor']
        sheet[f'J{r}'] = d['close_currency_factor']
        sheet[f'K{r}'] = d['currency_factor_ratio']
        sheet[f'L{r}'] = d['open_value_ILS_adjusted_forex']
        sheet[f'M{r}'] = d['close_date']
        sheet[f'N{r}'] = d['close_value_ILS']
        sheet[f'S{r}'] = d['position_type']
        sheet[f'T{r}'] = abs(d['quantity'])
        sheet[f'U{r}'] = d['open_price']
        sheet[f'V{r}'] = d['close_price']
        sheet[f'W{r}'] = abs(d['allocated_open_fee'])
        sheet[f'X{r}'] = abs(d['allocated_close_fee'])
        sheet[f'Y{r}'] = d['profit']
        ind_line += 1

        if 'profit_ILS_forex' not in d:
            continue
        value = d['profit_ILS_forex']
        if value >= 0:
            sheet[f'O{r}'] = value
            total_profit += value
        else:
            sheet[f'P{r}'] = value
            total_loss += value
        total_net  += value
        total_sell += d['gross_sale_value_ILS']

    return start_row + ind_line, total_net, total_sell, total_profit, total_loss


def _write_dividends_half(sheet, dividends_list, half_label, start_row):
    total_div, total_tax, ind_line = 0, 0, 0
    for d in dividends_list:
        if not _in_half(d['datetime'], half_label):
            continue
        r = start_row + ind_line
        sheet[f'B{r}'] = ind_line + 1
        sheet[f'C{r}'] = d['date']
        sheet[f'D{r}'] = d['ticker']
        sheet[f'E{r}'] = d['currency']
        sheet[f'F{r}'] = d['dividend']
        sheet[f'G{r}'] = d['withholding_tax']
        sheet[f'H{r}'] = d['currency_factor']
        sheet[f'I{r}'] = d['dividend_ILS']
        sheet[f'J{r}'] = d['withholding_tax_ILS']
        total_div += d['dividend_ILS']
        total_tax += abs(d['withholding_tax_ILS'])
        ind_line  += 1
    return start_row + ind_line, total_div, total_tax


def _write_unmatched_wht_table(sheet, unmatched_wht, start_row, col_start=2, col_end=10):
    PRIOR_FILL = PatternFill(start_color='6B3A1F', end_color='6B3A1F', fill_type='solid')
    TOTAL_FILL = PatternFill(start_color='F4EAE8', end_color='F4EAE8', fill_type='solid')
    COLS       = ['B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J']

    _style_header_row(sheet, start_row, col_start, col_end,
                      'מס שנוכה במקור והוחזר', PRIOR_FILL)
    sheet.merge_cells(start_row=start_row, start_column=2, end_row=start_row, end_column=3)

    total_ils  = 0
    data_start = start_row + 1
    for i, wht in enumerate(unmatched_wht):
        r = data_start + i
        sheet[f'B{r}'] = i + 1
        sheet[f'C{r}'] = wht['date']
        sheet[f'D{r}'] = wht['ticker']
        sheet[f'E{r}'] = wht['currency']
        sheet[f'G{r}'] = abs(wht['amount'])
        sheet[f'H{r}'] = wht['rate']
        sheet[f'J{r}'] = abs(round_half_up(wht['amount_ils']))
        for col in COLS:
            sheet[f'{col}{r}'].border = _THIN_BORDER
        total_ils += wht['amount_ils']

    total_row = data_start + len(unmatched_wht)
    sheet[f'B{total_row}'] = 'סה"כ מס שנוכה במקור'
    sheet[f'J{total_row}'] = -abs(round_half_up(total_ils))
    _style_totals_row(sheet, total_row, col_start, col_end, TOTAL_FILL)
    for col in COLS:
        sheet[f'{col}{total_row}'].border = _THIN_BORDER
    return total_row + 1


def _write_other_fees_sheet(xfile, other_fees_data):
    ws         = xfile['Fees & Interest Income']
    all_fees   = other_fees_data['commissions'] + other_fees_data['fees']
    all_debit  = other_fees_data['interest_debit']
    all_credit = other_fees_data['interest_credit']
    all_entries = sorted(all_fees + all_debit + all_credit,
                         key=lambda x: datetime.datetime.strptime(x['date'], '%d/%m/%Y'))

    bold_font = Font(bold=True)
    cols      = ['B', 'C', 'D', 'E', 'F', 'G', 'H', 'I']

    for i, entry in enumerate(all_entries):
        r = 6 + i
        ws[f'B{r}'] = i + 1
        ws[f'C{r}'] = entry['date']
        ws[f'D{r}'] = entry['category']
        desc = re.sub(r'^[A-Z]{3}\s+', '', entry['description'])
        desc = re.sub(r'\s+for\s+\w+-\d{4}$', '', desc)
        ws[f'E{r}'] = desc
        ws[f'F{r}'] = entry['currency']
        ws[f'G{r}'] = entry['amount']
        ws[f'H{r}'] = entry['rate']
        ws[f'I{r}'] = round_half_up(entry['amount_ils'])
        for col in cols:
            ws[f'{col}{r}'].border = _THIN_BORDER

    total_row = 6 + len(all_entries) + 1
    for offset, label, entries in [
        (0, 'סה"כ עמלות',              all_fees),
        (1, 'סה"כ ריבית חובה (Debit)', all_debit),
        (2, 'סה"כ ריבית זכות',         all_credit),
    ]:
        r = total_row + offset
        ws[f'B{r}'] = label
        ws[f'G{r}'] = abs(sum(e['amount']     for e in entries))
        ws[f'I{r}'] = abs(round_half_up(sum(e['amount_ils'] for e in entries)))
        for col in cols:
            ws[f'{col}{r}'].border = _THIN_BORDER
            ws[f'{col}{r}'].font   = bold_font


def _write_cg_summary_table(sheet, h1_profit, h1_loss, h1_sell, h2_profit, h2_loss, h2_sell):
    TITLE_FILL = PatternFill(start_color='1F3E6B', end_color='1F3E6B', fill_type='solid')
    H1_FILL    = PatternFill(start_color='E8F4EA', end_color='E8F4EA', fill_type='solid')
    H2_FILL    = PatternFill(start_color='E8EEF8', end_color='E8EEF8', fill_type='solid')
    YEAR_FILL  = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')
    TITLE_FONT = Font(bold=True, color='FFFFFF', size=10)
    DATA_FONT  = Font(bold=True, size=10)
    CENTER     = Alignment(horizontal='center', vertical='center')
    LEFT       = Alignment(horizontal='left',   vertical='center')
    RIGHT      = Alignment(horizontal='right',  vertical='center')

    sheet.column_dimensions['Q'].width = 26
    sheet.column_dimensions['R'].width = 16

    rows = [
        (5,  'סיכום',                        None,                                                          TITLE_FILL, TITLE_FONT),
        (6,  'רווח מחצית ראשונה',            round_half_up(h1_profit),                                      H1_FILL,    DATA_FONT),
        (7,  'הפסד מחצית ראשונה',            round_half_up(abs(h1_loss)),                                   H1_FILL,    DATA_FONT),
        (8,  'סכום מכירות מחצית ראשונה',     round_half_up(h1_sell),                                        H1_FILL,    DATA_FONT),
        (9,  'רווח מחצית שנייה',             round_half_up(h2_profit),                                      H2_FILL,    DATA_FONT),
        (10, 'הפסד מחצית שנייה',             round_half_up(abs(h2_loss)),                                   H2_FILL,    DATA_FONT),
        (11, 'סכום מכירות מחצית שנייה',      round_half_up(h2_sell),                                        H2_FILL,    DATA_FONT),
        (12, 'רווח כולל שנתי',               round_half_up(h1_profit + h2_profit),                          YEAR_FILL,  DATA_FONT),
        (13, '(56/256)הפסד כולל שנתי',      round_half_up(abs(h1_loss + h2_loss)),                         YEAR_FILL,  DATA_FONT),
        (14, 'מכירות שנתי',                  round_half_up(h1_sell + h2_sell),                              YEAR_FILL,  DATA_FONT),
        (15, 'רווח שנתי לאחר קיזוז הפסדים', round_half_up((h1_profit + h2_profit) + (h1_loss + h2_loss)), YEAR_FILL,  DATA_FONT),
    ]

    for row_num, label, value, fill, font in rows:
        sheet.row_dimensions[row_num].height = 18
        s = sheet.cell(row=row_num, column=17, value=label)
        s.fill, s.font, s.alignment = fill, font, (CENTER if value is None else LEFT)
        t = sheet.cell(row=row_num, column=18, value=value)
        t.fill, t.font, t.alignment = fill, font, RIGHT

def write_tax_form_files(file_dir, csv_file_name, closed_lots_list, inds_sorted_close_dates,
                         dividends_list, other_fees_data, unmatched_wht=None):
    template_file = os.path.dirname(os.path.abspath(__file__)) + '/tax_forms_template.xlsx'
    xfile = openpyxl.load_workbook(template_file)

    CG_COL_START,  CG_COL_END  = 1, 25
    DIV_COL_START, DIV_COL_END = 2, 10

    if closed_lots_list:
        sheet = xfile['Capital Gains (FOREX adjusted)']
        sheet.freeze_panes = 'A5'

        _style_header_row(sheet, 5, CG_COL_START, CG_COL_END, 'ינואר - יוני', _H1_FILL)
        h1_next, h1_net, h1_sell, h1_profit, h1_loss = _write_capital_gains_half(
            sheet, closed_lots_list, inds_sorted_close_dates, 'H1', 6)

        h2_header = h1_next + _TABLE_GAP
        _style_header_row(sheet, h2_header, CG_COL_START, CG_COL_END, 'יולי - דצמבר', _H2_FILL)
        h2_next, h2_net, h2_sell, h2_profit, h2_loss = _write_capital_gains_half(
            sheet, closed_lots_list, inds_sorted_close_dates, 'H2', h2_header + 1)

        _write_cg_summary_table(sheet, h1_profit, h1_loss, h1_sell, h2_profit, h2_loss, h2_sell)

    if dividends_list:
        sheet = xfile['Dividends']
        sheet.freeze_panes = 'A5'

        _style_header_row(sheet, 5, DIV_COL_START, DIV_COL_END, 'ינואר - יוני', _H1_FILL)
        h1_next, h1_div, h1_tax = _write_dividends_half(sheet, dividends_list, 'H1', 6)
        sheet[f'B{h1_next}'] = 'סה"כ מחצית ראשונה'
        sheet[f'I{h1_next}'] = h1_div
        sheet[f'J{h1_next}'] = h1_tax
        _style_totals_row(sheet, h1_next, DIV_COL_START, DIV_COL_END, _H1_TOTALS_FILL)

        h2_header = h1_next + _TABLE_GAP
        _style_header_row(sheet, h2_header, DIV_COL_START, DIV_COL_END, 'יולי - דצמבר', _H2_FILL)
        h2_next, h2_div, h2_tax = _write_dividends_half(sheet, dividends_list, 'H2', h2_header + 1)
        sheet[f'B{h2_next}'] = 'סה"כ מחצית שנייה'
        sheet[f'I{h2_next}'] = h2_div
        sheet[f'J{h2_next}'] = h2_tax
        _style_totals_row(sheet, h2_next, DIV_COL_START, DIV_COL_END, _H2_TOTALS_FILL)

        year_row = h2_next + 2
        sheet[f'B{year_row}'] = 'סה"כ שנתי'
        sheet[f'I{year_row}'] = round_half_up(h1_div + h2_div)
        sheet[f'J{year_row}'] = round_half_up(h1_tax + h2_tax)
        _style_totals_row(sheet, year_row, DIV_COL_START, DIV_COL_END,
                          PatternFill(start_color='BFBFBF', end_color='BFBFBF', fill_type='solid'))

        if unmatched_wht:
            _write_unmatched_wht_table(sheet, unmatched_wht, year_row + 2, DIV_COL_START, DIV_COL_END)

    _write_other_fees_sheet(xfile, other_fees_data)

    file_path = f'{file_dir}/tax_forms_{csv_file_name}.xlsx'
    xfile.save(file_path)
    print(f'Written: {file_path}')


# ---------------------------------------------------------------------------
# Entry point
# ---------------------------------------------------------------------------
def _extract_all(file_dir, csv_file_name, verbosity, date_slash_format):
    return (
        extract_trades_data_from_csv(file_dir, csv_file_name, verbosity=verbosity, date_slash_format=date_slash_format),
        extract_dividends_data_from_csv(file_dir, csv_file_name, verbosity=verbosity, date_slash_format=date_slash_format),
        extract_other_fees_data_from_csv(file_dir, csv_file_name, verbosity=verbosity, date_slash_format=date_slash_format),
    )


def generate_tax_forms(file_dir, csv_file_name, verbosity=0):
    try:
        trades, dividends_tuple, fees = _extract_all(file_dir, csv_file_name, verbosity, 'normal')
    except ValueError:
        print("*** failed to use date_slash_format='normal', attempting date_slash_format='USA'")
        trades, dividends_tuple, fees = _extract_all(file_dir, csv_file_name, verbosity, 'USA')

    closed_lots_list, inds_sorted_close_dates = trades
    dividends_list, unmatched_wht             = dividends_tuple
    write_tax_form_files(file_dir, csv_file_name, closed_lots_list, inds_sorted_close_dates,
                         dividends_list, fees, unmatched_wht=unmatched_wht)
    print('Finished generating tax forms.')


if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Run tax forms generator")
    parser.add_argument("-dir",           "--dir",       type=str, required=True)
    parser.add_argument("-csv_file_name", "--csv_name",  type=str, required=True)
    parser.add_argument("-verbosity",     "--verbosity",  type=int, required=False, default=0)
    args = parser.parse_args()
    generate_tax_forms(args.dir, args.csv_name, args.verbosity)